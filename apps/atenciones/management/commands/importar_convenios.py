"""
Importa tarifas de convenios desde el Excel de contratación.

Uso:
    python manage.py importar_convenios "VALORES CONTRATACION A MAYO 2026.xlsx"

Lee la hoja 'contratacion' (Id, Nit, Entidad, ServicioId, CodigoServicio,
Servicio, TipoContratacion, Porcentaje, ValorContratado, Observacion),
crea/actualiza Entidades (con NIT) y hace upsert de TarifaConvenio por
(entidad, código de servicio). Re-ejecutable: actualiza valores existentes.
"""

from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.atenciones.models import Entidad, TarifaConvenio, normalizar


def _decimal(valor) -> Decimal:
    try:
        return Decimal(str(valor).replace(",", ".").strip() or "0")
    except (InvalidOperation, AttributeError):
        return Decimal("0")


class Command(BaseCommand):
    help = "Importa tarifas de convenios (hoja 'contratacion')."

    def add_arguments(self, parser):
        parser.add_argument("ruta", help="Ruta del xlsx de contratación")
        parser.add_argument("--hoja", default="contratacion")

    @transaction.atomic
    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["ruta"], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No existe el archivo: {opts['ruta']}")
        if opts["hoja"] not in wb.sheetnames:
            raise CommandError(f"No existe la hoja {opts['hoja']!r}. Hojas: {wb.sheetnames}")

        ws = wb[opts["hoja"]]
        ws.reset_dimensions()
        filas = ws.iter_rows(values_only=True)
        encabezado = [normalizar(h) for h in next(filas)]
        idx = {nombre: i for i, nombre in enumerate(encabezado)}
        for col in ("NIT", "ENTIDAD", "CODIGOSERVICIO", "VALORCONTRATADO"):
            if col not in idx:
                raise CommandError(f"Falta la columna {col} en la hoja. Encabezado: {encabezado}")

        entidades: dict[str, Entidad] = {
            e.nombre_normalizado: e for e in Entidad.objects.all()
        }
        creadas = tarifas_nuevas = tarifas_actualizadas = omitidas = 0

        for fila in filas:
            def val(col):
                i = idx.get(col)
                return fila[i] if i is not None and i < len(fila) else None

            nombre = str(val("ENTIDAD") or "").strip()
            codigo = str(val("CODIGOSERVICIO") or "").strip()
            if not nombre or not codigo:
                omitidas += 1
                continue

            clave = normalizar(nombre)
            entidad = entidades.get(clave)
            if entidad is None:
                entidad = Entidad.objects.create(
                    nombre=nombre, nit=str(val("NIT") or "").strip()
                )
                entidades[clave] = entidad
                creadas += 1
            elif not entidad.nit and val("NIT"):
                entidad.nit = str(val("NIT")).strip()
                entidad.save(update_fields=["nit"])

            _, creada = TarifaConvenio.objects.update_or_create(
                entidad=entidad,
                codigo_servicio=codigo,
                defaults={
                    "nombre_servicio": str(val("SERVICIO") or "").strip()[:200],
                    "tipo_contratacion": str(val("TIPOCONTRATACION") or "").strip()[:60],
                    "porcentaje": _decimal(val("PORCENTAJE")),
                    "valor": _decimal(val("VALORCONTRATADO")),
                },
            )
            if creada:
                tarifas_nuevas += 1
            else:
                tarifas_actualizadas += 1

        wb.close()
        self.stdout.write(self.style.SUCCESS(
            f"Convenios OK. Entidades nuevas: {creadas} | tarifas nuevas: {tarifas_nuevas} | "
            f"actualizadas: {tarifas_actualizadas} | filas omitidas: {omitidas}"
        ))
