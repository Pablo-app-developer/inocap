"""
Importa el mapeo Especialidad → Unidad de negocio (hoja 'Unidades' del Excel
de Medicloud; es el VLOOKUP que agregaba la columna 'Unidad de Negocio').

Uso:
    python manage.py importar_mapeo_unidades "2026 MEDICLOUD.xlsx" --hoja Unidades

Crea las Unidades de negocio que no existan (match tolerante a tildes y
mayúsculas, ej. 'Clínica de Alergías' ≡ 'Clínica de Alergias' existente).
"""

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from apps.atenciones.models import MapeoEspecialidad, normalizar
from apps.core.models import UnidadNegocio


class Command(BaseCommand):
    help = "Importa el mapeo Especialidad → Unidad de negocio."

    def add_arguments(self, parser):
        parser.add_argument("ruta")
        parser.add_argument("--hoja", default="Unidades")

    @transaction.atomic
    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["ruta"], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No existe el archivo: {opts['ruta']}")
        if opts["hoja"] not in wb.sheetnames:
            raise CommandError(f"No existe la hoja {opts['hoja']!r}. Hojas: {wb.sheetnames}")

        ws = wb[opts["hoja"]]
        ws.reset_dimensions()
        filas = ws.iter_rows(values_only=True)
        next(filas)  # encabezado: Especialidad | UNIDAD DE NEGOCIO

        unidades = {normalizar(u.nombre): u for u in UnidadNegocio.objects.all()}
        mapeos = unidades_creadas = 0

        for fila in filas:
            if not fila or not fila[0] or len(fila) < 2 or not fila[1]:
                continue
            especialidad = str(fila[0]).strip()
            nombre_unidad = str(fila[1]).strip()

            clave = normalizar(nombre_unidad)
            unidad = unidades.get(clave)
            if unidad is None:
                unidad = UnidadNegocio.objects.create(
                    nombre=nombre_unidad, codigo=slugify(nombre_unidad)
                )
                unidades[clave] = unidad
                unidades_creadas += 1

            MapeoEspecialidad.objects.update_or_create(
                especialidad_normalizada=normalizar(especialidad),
                defaults={"especialidad": especialidad, "unidad_negocio": unidad},
            )
            mapeos += 1

        wb.close()
        self.stdout.write(self.style.SUCCESS(
            f"Mapeos cargados: {mapeos} | unidades de negocio creadas: {unidades_creadas}"
        ))
