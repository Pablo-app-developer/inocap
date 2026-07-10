"""
Importa atenciones de Medicloud desde un xlsx (export manual o de la API).

Uso:
    python manage.py importar_atenciones "2026 MEDICLOUD.xlsx" --hoja BASE --reemplazar-anio 2026
    python manage.py importar_atenciones descarga_api.xlsx --hoja indicadores

Columnas usadas (por nombre, tolerante a orden): AG D, TC Codi, TC Nomb,
AG Asistenc, EN Codi, EN Nomb, Valor Servicio, Sede, Sala, Especialidad,
y 'Unidad de Negocio' si existe (el export de la API no la trae: en ese caso
se deriva de Especialidad vía MapeoEspecialidad).

NO importa datos personales (nombres, documentos, teléfonos, diagnósticos).

Antes de importar por primera vez, carga el mapeo de especialidades:
    python manage.py importar_mapeo_unidades "2026 MEDICLOUD.xlsx" --hoja Unidades
"""

from __future__ import annotations

import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from apps.atenciones.models import Atencion, Entidad, MapeoEspecialidad, normalizar
from apps.core.models import UnidadNegocio

LOTE = 5000


def _fecha(valor):
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    s = str(valor or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _decimal(valor):
    try:
        return Decimal(str(valor).replace(",", ".").strip() or "0")
    except (InvalidOperation, AttributeError):
        return Decimal("0")


class Command(BaseCommand):
    help = "Importa atenciones desde un export de Medicloud (xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("ruta")
        parser.add_argument("--hoja", default="BASE")
        parser.add_argument(
            "--reemplazar-anio", type=int, default=None,
            help="Borra las atenciones existentes de ese año antes de importar.",
        )

    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["ruta"], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"No existe el archivo: {opts['ruta']}")
        if opts["hoja"] not in wb.sheetnames:
            raise CommandError(f"No existe la hoja {opts['hoja']!r}. Hojas: {wb.sheetnames}")

        ws = wb[opts["hoja"]]
        ws.reset_dimensions()  # los xlsx de la API declaran dimensiones incorrectas
        filas = ws.iter_rows(values_only=True)
        encabezado = [str(h).strip() if h is not None else "" for h in next(filas)]
        idx = {nombre: i for i, nombre in enumerate(encabezado)}

        requeridas = ["AG D", "TC Codi", "TC Nomb", "AG Asistenc", "EN Codi",
                      "EN Nomb", "Valor Servicio", "Sede", "Sala", "Especialidad"]
        faltantes = [c for c in requeridas if c not in idx]
        if faltantes:
            raise CommandError(f"Faltan columnas: {faltantes}")
        col_unidad = idx.get("Unidad de Negocio")

        # Cachés en memoria para no consultar por fila.
        entidades = {e.nombre_normalizado: e for e in Entidad.objects.all()}
        unidades = {normalizar(u.nombre): u for u in UnidadNegocio.objects.all()}
        mapeo_esp = {
            m.especialidad_normalizada: m.unidad_negocio
            for m in MapeoEspecialidad.objects.select_related("unidad_negocio")
        }

        def entidad_de(nombre_crudo):
            nombre = str(nombre_crudo or "").strip()
            if not nombre:
                return None
            clave = normalizar(nombre)
            e = entidades.get(clave)
            if e is None:
                e = Entidad.objects.create(nombre=nombre, creada_automaticamente=True)
                entidades[clave] = e
            return e

        def unidad_de(fila):
            # 1) columna 'Unidad de Negocio' del propio archivo (export manual)
            if col_unidad is not None and col_unidad < len(fila) and fila[col_unidad]:
                u = unidades.get(normalizar(fila[col_unidad]))
                if u is not None:
                    return u
            # 2) derivada de Especialidad (caso API)
            esp = fila[idx["Especialidad"]] if idx["Especialidad"] < len(fila) else None
            return mapeo_esp.get(normalizar(esp)) if esp else None

        if opts["reemplazar_anio"]:
            borradas, _ = Atencion.objects.filter(anio=opts["reemplazar_anio"]).delete()
            self.stdout.write(f"Atenciones {opts['reemplazar_anio']} borradas: {borradas}")

        lote: list[Atencion] = []
        importadas = sin_fecha = sin_unidad = 0

        def volcar():
            nonlocal lote
            if lote:
                with transaction.atomic():
                    Atencion.objects.bulk_create(lote)
                lote = []

        for fila in filas:
            def val(col):
                i = idx[col]
                return fila[i] if i < len(fila) else None

            fecha = _fecha(val("AG D"))
            if fecha is None:
                sin_fecha += 1
                continue

            unidad = unidad_de(fila)
            if unidad is None:
                sin_unidad += 1  # se importa igual (unidad NULL) para poder auditar

            lote.append(Atencion(
                fecha=fecha, anio=fecha.year, mes=fecha.month,
                codigo_servicio=str(val("TC Codi") or "").strip()[:30],
                nombre_servicio=str(val("TC Nomb") or "").strip()[:200],
                estado=str(val("AG Asistenc") or "").strip()[:30],
                entidad=entidad_de(val("EN Nomb")),
                entidad_codigo=str(val("EN Codi") or "").strip()[:30],
                unidad_negocio=unidad,
                especialidad=str(val("Especialidad") or "").strip()[:120],
                sede_nombre=str(val("Sede") or "").strip()[:120],
                sala_nombre=str(val("Sala") or "").strip()[:120],
                valor_servicio=_decimal(val("Valor Servicio")),
            ))
            importadas += 1
            if len(lote) >= LOTE:
                volcar()
                if importadas % 50000 == 0:
                    self.stdout.write(f"  ... {importadas} filas")

        volcar()
        wb.close()
        self.stdout.write(self.style.SUCCESS(
            f"Atenciones importadas: {importadas} | sin fecha (omitidas): {sin_fecha} | "
            f"sin unidad de negocio (importadas con unidad NULL): {sin_unidad}"
        ))
